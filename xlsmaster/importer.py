import openpyxl


class XlsUtilities:
    def __init__(self, filepath):
        self.filepath = filepath
        self.book = openpyxl.load_workbook(filepath)

    def reload(self):
        self.book = openpyxl.load_workbook(self.filepath)

    def switch_to_active(self):
        return self.book.active

    def get_cell_by_coord(self, row, col):
        sheet = self.switch_to_active()
        return sheet.cell(row=row, column=col)

    def get_cell_by_position(self,cell):
        sheet = self.switch_to_active()
        return sheet[cell]

    def get_max_row(self):
        return self.switch_to_active().max_row

    def get_max_column(self):
        return self.switch_to_active().max_column

    def get_specific_row_values(self, row):
        max_col = self.get_max_column()
        results = list()
        for col in range(1, max_col + 1):
            results.append(self.get_cell_by_coord(row, col).value)
        return results

    def get_specific_column_values(self, col):
        max_row = self.get_max_row()
        results = list()
        for row in range(1, max_row + 1):
            results.append(self.get_cell_by_coord(row, col).value)
        return results
    def get_specific_column_values_by_value(self,value):
        data = self.iterate()
        select = list()
        coord = -1
        for subdata in data[0]:
            if value == subdata:
                coord = data[0].index(value)
                break
        for idx in data[1:]:
            select.append(idx[coord])
        return select

    def get_specific_row_values_by_value(self,value):
        data = self.iterate()
        select = list()
        for idx in data:
            if value == idx[0]:
                select = idx[1:]
                break
        return select

    def iterate(self):
        max_col = self.get_max_column()
        max_row = self.get_max_row()
        final_dataset = list()
        dataset = list()
        for row in range(1, max_row+1):
            temp = list()
            for col in range(1, max_col+1):
                temp.append(self.get_cell_by_coord(row,col).value)
            dataset.append(temp)
        header_is_set = False
        template = {}
        for data in dataset:
            temp = {}
            if not header_is_set:
                for idx in range(0,len(data)):
                    template[data[idx]] = ""
                    header_is_set = True
                continue
            for k in template.keys():
                temp[k] = data[list(template.keys()).index(k)]
                idx+=1
            final_dataset.append(temp)

        return final_dataset